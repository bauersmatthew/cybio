#!/usr/bin/env python

# Script: "cb-np_count-isoforms.py"
# Requires: Python 3, Scipy [, openpyxl]
# Purpose: This script, given the output of cb-np_count-isoforms.py, will
#          calculate the pairwise correlations for each pair of features.
# Author: Matthew Bauer

import argparse
import sys
import os.path
import re
from scipy import stats

def get_all_matches(l, expr):
    """Get all full REGEX matches to expr in l."""
    ret = []
    for e in l:
        m = re.match(expr, e)
        if m is not None and m.start() == 0 and m.end() == len(e):
            ret.append(e)
    return ret

def valid_file(p):
    """Raise an exception if p is not a valid file; otherwise, return p."""
    if not os.path.isfile(p):
        raise RuntimeError("'{}' is not a valid input file!".format(p))
    return p

arg_parser = argparse.ArgumentParser(
    description=('Calculate pairwise correlations for each pair of features.'),
    epilog='Results are written to STDOUT.')
arg_parser.add_argument('-V', '--version', action='version',
                        version='%(prog)s 1.0.0')
arg_parser.add_argument('isoforms',
                        help='The output file from cb-np_count-isoforms.py.',
                        type=valid_file)
arg_parser.add_argument('-t', '--test',
                        action='append', type=str, metavar='test_pair',
                        default=[],
                        help=('A description of pairs to calculate '
                              'correlations for. Format: first|second. '
                              'This argument can be passed more '
                              'than once to describe multiple test pairs.'))
arg_parser.add_argument('-T', '--test-regex',
                        action='append', type=str, metavar='test_pair',
                        default=[],
                        help=('The same as -t/--test, but the first and second '
                              'strings are interpreted as REGEX expressions.'))
arg_parser.add_argument('-x', '--xlsx',
                        action='store', type=str, metavar='output.xlsx',
                        help=('Write output to an XLSX file. TSV output will '
                              'be written to STDOUT.'))
arg_parser.add_argument('-d', '--delim',
                        action='store', type=str, default='|',
                        help=('Set the delimiter that should be used to '
                              'separate feature names in test strings. '
                              'Default: |'))
arg_parser.add_argument('-i', '--allow-identity-tests',
                        action='store_true', default=False,
                        help=('Allow identity tests e.g. \'F1|F1\' to be '
                              ' performed. Default: ignore them.'))
arg_parser.add_argument('--count-only',
                        action='store_true', default=False,
                        help=('Do not do any chi-squared analysis; simply '
                              'output the counts (observed values).'))
args = arg_parser.parse_args()

# read isoform table; construct set of all features
iso_counts = []
features = set()
with open(args.isoforms) as fin:
    for line in fin:
        line = line.rstrip()
        if line:
            fields = line.split('\t')
            f = fields[0].split(',')
            c = int(fields[1])
            iso_counts.append((f, c))
            for feat in f:
                features.add(feat)
features = list(sorted(features))

# collect features of interest
tests = set()
def parse_teststr(t, use_regex):
    sides = t.split(args.delim)
    if len(sides) != 2:
        sys.stderr.write('Improper test: \'{}\''.format(t))
        sys.exit(-1)
    if use_regex:
        s0_matches = get_all_matches(features, sides[0])
        s1_matches = get_all_matches(features, sides[1])
        for s0m in s0_matches:
            for s1m in s1_matches:
                if ((args.allow_identity_tests or s0m != s1m)
                    and (s1m, s0m) not in tests):
                    tests.add((s0m, s1m))
    elif ((args.allow_identity_tests or sides[0] != sides[1])
          and (sides[1], sides[0]) not in tests):
        tests.add(tuple(sides))

for t in args.test:
    parse_teststr(t, False)
for t in args.test_regex:
    parse_teststr(t, True)

def get_isos_with_and_without(f, counts):
    """Get isoform count lines with and without the given feature."""
    isos_with = []
    isos_without = []
    for c in counts:
        if f in c[0]:
            isos_with.append(c)
        else:
            isos_without.append(c)
    return isos_with, isos_without

def do_test(t):
    """Conduct one test."""
    yes0, no0 = get_isos_with_and_without(t[0], iso_counts)
    yes0yes1, yes0no1 = get_isos_with_and_without(t[1], yes0)
    no0yes1, no0no1 = get_isos_with_and_without(t[1], no0)

    # conduct chi-squared
    #    | y0 n0 |
    # ---+-------+
    # y1 | yy ny |
    # n1 | yn nn |
    # ---+-------+
    # (look at wiki)

    # observed values
    count = lambda l: sum([x[1] for x in l])
    o_yy = count(yes0yes1)
    o_yn = count(yes0no1)
    o_ny = count(no0yes1)
    o_nn = count(no0no1)

    if args.count_only:
        return o_yy, o_yn, o_ny, o_nn

    # row/column totals
    tot = o_yy+o_yn+o_ny+o_nn
    r1t = o_yy+o_ny
    r2t = o_yn+o_nn
    c1t = o_yy+o_yn
    c2t = o_ny+o_nn
    # expected values
    e_yy = c1t*r1t/tot
    e_ny = c2t*r1t/tot
    e_yn = c1t*r2t/tot
    e_nn = c2t*r2t/tot

    # chi-squared partials
    x2_yy = pow(o_yy-e_yy, 2)/e_yy
    x2_ny = pow(o_ny-e_ny, 2)/e_ny
    x2_yn = pow(o_yn-e_yn, 2)/e_yn
    x2_nn = pow(o_nn-e_nn, 2)/e_nn
    x2 = x2_yy+x2_ny+x2_yn+x2_nn
    # compute p-value
    p = 1-stats.chi2.cdf(x2, 1)
    return o_yy, o_yn, o_ny, o_nn, x2, p

# conduct tests, print results
sys.stdout.write('first\tsecond\tyes-yes\tyes-no\tno-yes\tno-no')
if not args.count_only:
    sys.stdout.write('\tchi-sq\tp-value')
sys.stdout.write('\n')
xlsx_wb = None
pvals = []
if args.xlsx is not None:
    from openpyxl import Workbook
    xlsx_wb = Workbook()
for t in sorted(sorted(tests, key=lambda x: x[1]), key=lambda y: y[0]):
    inf = do_test(t)
    if not args.count_only:
        sys.stdout.write('{}\t{}\t{}\t{}\t{}\t{}\t{}\t{}\n'.format(t[0], t[1],
                                                                   *inf))
    else:
        sys.stdout.write('{}\t{}\t{}\t{}\t{}\t{}\n'.format(t[0], t[0], *inf))
    if xlsx_wb is not None:
        import openpyxl
        o_yy, o_yn, o_ny, o_nn = inf[:4]
        if not args.count_only:
            x2, p = inf[4:]
        ws = xlsx_wb.create_sheet('{} vs {}'.format(*t))
        ws['A1'] = 'Feature 1'
        ws['B1'] = t[0]
        ws['A2'] = 'Feature 2'
        ws['B2'] = t[1]
        ws['A4'] = 'Data'
        ws['C4'] = '{} present?'.format(t[0])
        ws.merge_cells('C4:D4')
        alig = openpyxl.styles.Alignment(horizontal='center',
                                         vertical='center')
        ws['C4'].alignment = alig
        ws['C5'] = 'Yes'
        ws['D5'] = 'No'
        ws['A6'] = '{} present?'.format(t[1])
        ws.merge_cells('A6:A7')
        ws['A6'].alignment = alig
        ws['B6'] = 'Yes'
        ws['B7'] = 'No'
        ws['C6'] = o_yy
        ws['D6'] = o_ny
        ws['C7'] = o_yn
        ws['D7'] = o_nn

        if not args.count_only:
            ws['A10'] = '\u03c7\u00b2' # X2
            ws['B10'] = x2
            ws['A11'] = 'P-val'
            ws['B11'] = p
            import openpyxl
            fill = None
            if p < 0.01:
                fill = openpyxl.styles.colors.GREEN
            elif p < 0.05:
                fill = openpyxl.styles.colors.YELLOW
            else:
                fill = openpyxl.styles.colors.RED
            ws['B11'].fill = openpyxl.styles.fills.PatternFill(
                patternType='solid',
                fgColor=fill)
            pvals.append((ws.title, p))

if xlsx_wb is not None:
    ws = xlsx_wb['Sheet']
    ws.title = 'Overview'
    if args.count_only:
        ws['A1'] = 'P-values were not calculated.'
    else:
        ws['A1'] = 'Pair'
        ws['B1'] = 'P-val'
        import openpyxl
        for row in range(len(pvals)):
            ws.cell(column=1, row=2+row).value = pvals[row][0]
            ws.cell(column=2, row=2+row).value = pvals[row][1]
            f = openpyxl.styles.colors.RED
            if pvals[row][1] < 0.01:
                f = openpyxl.styles.colors.GREEN
            elif pvals[row][1] < 0.05:
                f = openpyxl.styles.colors.YELLOW
            ws.cell(column=2, row=2+row).fill = \
                openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=f)
    xlsx_wb.save(args.xlsx)
